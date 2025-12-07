import streamlit as st
import sqlite3
import pandas as pd
import plotly.express as px
from datetime import date, time as dt_time, datetime
import io
import time
import re
import os

# --- ИМПОРТЫ ДЛЯ EXCEL ---
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

DB_FILE = 'construction_log.db'

# --- CSS ---
def local_css():
    st.markdown("""
    <style>
        :root { --primary-blue: #003366; --accent-green: #8DB600; --text-grey: #333333; --bg-light: #FFFFFF; --input-bg: #F0F2F6; }
        .stApp { background-color: var(--bg-light); color: var(--text-grey); }
        h1, h2, h3, h4, h5, h6, p, label, li, span, div { color: var(--text-grey); }
        .logo-container { font-family: 'Arial', sans-serif; font-weight: bold; font-size: 40px; line-height: 1; }
        .logo-pro { color: var(--primary-blue) !important; }
        .logo-maintain { color: var(--accent-green) !important; }
        .slogan { font-size: 12px; color: var(--primary-blue) !important; margin-top: 5px; font-weight: bold; }
        .phone-header { text-align: right; color: var(--primary-blue) !important; font-size: 20px; font-weight: bold; padding-top: 10px; }
        div.stButton > button { background-color: var(--primary-blue) !important; color: #FFFFFF !important; border-radius: 4px; border: none; font-weight: bold; }
        div.stButton > button:hover { background-color: var(--accent-green) !important; color: #FFFFFF !important; }
        div.stButton > button p { color: #FFFFFF !important; }
        input[type="text"], input[type="number"], .stDateInput input { background-color: var(--input-bg) !important; color: #000000 !important; border: 1px solid #ccc; }
        div[data-baseweb="select"] > div { background-color: var(--input-bg) !important; border-color: #ccc !important; }
        div[data-baseweb="select"] span { color: #000000 !important; }
        ul[data-baseweb="menu"] { background-color: #FFFFFF !important; }
        ul[data-baseweb="menu"] li span { color: var(--text-grey) !important; }
        section[data-testid="stSidebar"] { background-color: #F8F9FB; }
        section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2 { color: var(--primary-blue) !important; }
        div[data-testid="stDataFrame"] { border: 1px solid #ddd; }
        .footer { position: fixed; left: 0; bottom: 0; width: 100%; background-color: #f1f1f1; color: #555555 !important; text-align: center; font-size: 12px; padding: 10px; border-top: 1px solid #ddd; z-index: 100; }
        .block-container { padding-bottom: 50px; }
    </style>
    """, unsafe_allow_html=True)

# --- DB INIT ---
def init_db(force_reset=False):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    if force_reset:
        c.execute("DROP TABLE IF EXISTS work_logs")
        c.execute("DROP TABLE IF EXISTS scaffolds")
        c.execute("DROP TABLE IF EXISTS projects")
        c.execute("DROP TABLE IF EXISTS workers")
        
    c.execute('''CREATE TABLE IF NOT EXISTS projects (id INTEGER PRIMARY KEY, name TEXT UNIQUE)''')
    c.execute('''CREATE TABLE IF NOT EXISTS workers (id INTEGER PRIMARY KEY, name TEXT UNIQUE, position TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS scaffolds (
        id INTEGER PRIMARY KEY, 
        project_id INTEGER NOT NULL, 
        number TEXT, 
        description TEXT, 
        volume_m3 REAL, 
        area_m2 REAL, 
        weight_to REAL, 
        material_cost REAL, 
        acc TEXT, 
        FOREIGN KEY(project_id) REFERENCES projects(id), 
        UNIQUE(project_id, number))''')
    c.execute('''CREATE TABLE IF NOT EXISTS work_logs (id INTEGER PRIMARY KEY, user_name TEXT, project_name TEXT, scaffold_number TEXT, work_date DATE, hours REAL, comment TEXT, version TEXT)''')
    conn.commit()
    conn.close()

if not os.path.exists(DB_FILE):
    init_db()

# --- STATE ---
if 'logged_in' not in st.session_state: st.session_state['logged_in'] = False
if 'user_role' not in st.session_state: st.session_state['user_role'] = None 
if 'current_user_name' not in st.session_state: st.session_state['current_user_name'] = None
if 'admin_warning_shown' not in st.session_state: st.session_state['admin_warning_shown'] = False

def get_data(query, params=()):
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def run_query(query, params=()):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute(query, params)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        return str(e)

# --- HELPER FUNCTIONS ---
def safe_float(val):
    if pd.isna(val) or val == '': return 0.0
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        try: return float(val.replace(',', '.').replace(' ', '').strip())
        except: return 0.0
    return 0.0

def parse_hours(val):
    if pd.isna(val) or val == '': return 0.0
    if isinstance(val, dt_time): return val.hour + val.minute / 60.0
    if isinstance(val, datetime): return val.hour + val.minute / 60.0
    return safe_float(val)

def clean_scaffold_number(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    if s.endswith(".0"): return s[:-2]
    return s

def get_col_val(row, possibilities):
    for col in possibilities:
        if col in row: return row[col]
    return None

def get_export_filename(selected_projects):
    if selected_projects and len(selected_projects) == 1:
        proj_name = selected_projects[0]
        match = re.match(r'^([\d-]+)', proj_name)
        if match: return f"{match.group(1)}_Engineering Stunden.xlsx"
        else:
            safe_name = "".join([c for c in proj_name if c.isalnum() or c in (' ', '-', '_')]).strip()[:15]
            return f"{safe_name}_Engineering Stunden.xlsx"
    else:
        return "Gesamt_Engineering Stunden.xlsx"

def to_excel(df, sheet_name='Report'):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)): cell.alignment = Alignment(horizontal='right')
                else: cell.alignment = Alignment(horizontal='left')
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column_letter].width = adjusted_width
    return output.getvalue()

def get_template_excel():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_g = pd.DataFrame(columns=['Gerüstnummer', 'Beschreibung', 'm3', 'm2', 'to', 'Materialwert', 'ACC'])
        df_g.to_excel(writer, index=False, sheet_name='Gerüste')
        df_s = pd.DataFrame(columns=['Datum', 'Name', 'Gerüstnummer', 'Stunden', 'Anmerkungen', 'Versionsnummer'])
        df_s.to_excel(writer, index=False, sheet_name='Stundenübersicht')
    return output.getvalue()

def logout():
    st.session_state['logged_in'] = False
    st.session_state['user_role'] = None; st.rerun()

def render_header():
    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown("""<div class="logo-container"><span class="logo-pro">pro</span><span class="logo-maintain">maintain</span><sup>®</sup></div><div class="slogan">Nah am Kunden. Nah am Projekt.</div>""", unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="phone-header">📞 08458 320800</div>', unsafe_allow_html=True)
    st.markdown("---")

st.set_page_config(page_title="promaintain Portal", page_icon="🏗️", layout="wide")
local_css()

# ================= LOGIN =================
if not st.session_state['logged_in']:
    render_header()
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.subheader("System-Login")
        workers_df = get_data("SELECT name FROM workers ORDER BY name")
        workers_list = workers_df['name'].tolist()
        
        selected_worker = st.selectbox("Mitarbeiter auswählen", workers_list if workers_list else [""])
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Als Mitarbeiter anmelden", use_container_width=True):
            if workers_list:
                st.session_state['logged_in'] = True
                st.session_state['user_role'] = 'worker'
                st.session_state['current_user_name'] = selected_worker
                st.rerun()

        st.markdown("---")
        if not st.session_state['admin_warning_shown']:
            if st.button("Administrator-Login", use_container_width=True):
                st.session_state['admin_warning_shown'] = True
                st.rerun()
        else:
            st.warning("⚠️ Zugriff nur für Projektmanager!")
            c1, c2 = st.columns(2)
            if c1.button("OK", type="primary"):
                st.session_state['logged_in'] = True; st.session_state['user_role'] = 'admin'; st.rerun()
            if c2.button("Cancel"): st.session_state['admin_warning_shown'] = False; st.rerun()

# ================= APP =================
else:
    if st.session_state['user_role'] == 'admin': st.sidebar.markdown("### 👤 Administrator")
    else: st.sidebar.markdown(f"### 👷 {st.session_state['current_user_name']}")
    if st.sidebar.button("Abmelden"): logout()
    render_header()

    # --- WORKER ---
    if st.session_state['user_role'] == 'worker':
        st.title(f"Willkommen, {st.session_state['current_user_name']}")
        tab_work, tab_design = st.tabs(["🕒 Arbeitszeit erfassen", "🏗️ Daten nach Planung"])

        with tab_work:
            st.subheader("📝 Stunden buchen (Stundenübersicht)")
            projects_df = get_data("SELECT id, name FROM projects")
            if not projects_df.empty:
                p_map = dict(zip(projects_df['name'], projects_df['id']))
                sel_proj = st.selectbox("Projekt wählen", projects_df['name'], key="w_proj")
                scaf_df = get_data("SELECT number, description FROM scaffolds WHERE project_id = ?", (p_map[sel_proj],))
                
                if not scaf_df.empty:
                    scaf_df['disp'] = scaf_df.apply(lambda x: f"{x['number']} ({x['description']})" if x['description'] else x['number'], axis=1)
                    with st.form("wf"):
                        s_scaf = st.selectbox("Gerüstnummer", scaf_df['disp'])
                        c1, c2 = st.columns(2)
                        w_date = c1.date_input("Datum", date.today())
                        hours = c2.number_input("Stunden", min_value=0.5, step=0.5)
                        c3, c4 = st.columns(2)
                        comment = c3.text_input("Anmerkungen (Optional)")
                        version = c4.text_input("Versionsnummer (Optional)")
                        if st.form_submit_button("Zeit buchen"):
                            run_query("INSERT INTO work_logs (user_name, project_name, scaffold_number, work_date, hours, comment, version) VALUES (?, ?, ?, ?, ?, ?, ?)", 
                                      (st.session_state['current_user_name'], sel_proj, s_scaf.split(" (")[0], w_date, hours, comment, version))
                            st.success("Gespeichert!"); time.sleep(1); st.rerun()
                else: st.info("Keine Gerüste für dieses Projekt.")
            else: st.warning("Keine Projekte.")

        with tab_design:
            st.subheader("🏗️ Daten nach abgeschlossener Planung erfassen")
            projects_df = get_data("SELECT id, name FROM projects")
            if not projects_df.empty:
                p_map = dict(zip(projects_df['name'], projects_df['id']))
                design_proj = st.selectbox("Ziel-Projekt", projects_df['name'], key="d_proj")
                scaf_df_design = get_data("SELECT id, number, description, acc, volume_m3, area_m2, weight_to, material_cost FROM scaffolds WHERE project_id = ?", (p_map[design_proj],))
                
                if not scaf_df_design.empty:
                    scaf_df_design['disp'] = scaf_df_design.apply(lambda x: f"{x['number']} ({x['description']})", axis=1)
                    selected_scaf_display = st.selectbox("Gerüst auswählen", scaf_df_design['disp'], key="d_scaf_sel")
                    sel_num = selected_scaf_display.split(" (")[0]
                    current_row = scaf_df_design[scaf_df_design['number'] == sel_num].iloc[0]
                    st.info(f"Daten für: **{sel_num}**")
                    with st.form("design_update_form"):
                        c1, c2 = st.columns(2)
                        new_desc = c1.text_input("Beschreibung", value=current_row['description'] if current_row['description'] else "")
                        acc_opts = ["", "ja", "nein"]
                        curr_acc = current_row['acc'] if current_row['acc'] in acc_opts else ""
                        new_acc = c2.selectbox("ACC", acc_opts, index=acc_opts.index(curr_acc))
                        c3, c4 = st.columns(2)
                        new_vol = c3.number_input("m³ (Volumen)", min_value=0.0, step=1.0, value=float(current_row['volume_m3']) if current_row['volume_m3'] else 0.0)
                        new_area = c4.number_input("m² (Fläche)", min_value=0.0, step=1.0, value=float(current_row['area_m2']) if current_row['area_m2'] else 0.0)
                        c5, c6 = st.columns(2)
                        new_weight = c5.number_input("to (Gewicht) *Pflichtfeld", min_value=0.0, step=0.1, value=float(current_row['weight_to']) if current_row['weight_to'] else 0.0)
                        new_cost = c6.number_input("Materialwert (€) *Pflichtfeld", min_value=0.0, step=100.0, value=float(current_row['material_cost']) if current_row['material_cost'] else 0.0)
                        if st.form_submit_button("Speichern"):
                            if new_weight > 0 and new_cost > 0:
                                success = run_query("UPDATE scaffolds SET description=?, acc=?, volume_m3=?, area_m2=?, weight_to=?, material_cost=? WHERE project_id=? AND number=?", 
                                                    (new_desc, new_acc, new_vol, new_area, new_weight, new_cost, p_map[design_proj], sel_num))
                                if success == True: st.success("Gespeichert!"); time.sleep(1.5); st.rerun()
                            else: st.error("Fehler: 'to' und 'Materialwert' müssen > 0 sein.")
                else: st.warning("Keine Gerüste.")
            else: st.warning("Keine Projekte.")

    # --- ADMIN ---
    elif st.session_state['user_role'] == 'admin':
        st.title("⚙️ Administrationsbereich")
        tab0, tab1, tab2, tab3, tab4 = st.tabs(["📋 Gerüstübersicht", "📈 KPI & Analytik", "👥 Mitarbeiter", "🏗️ Projekte/Gerüste", "📥 Daten-Import"])

        # TAB 0: MASTER TABLE
        with tab0:
            st.subheader("📋 Gerüstübersicht (Master-Tabelle)")
            query_raw = "SELECT * FROM work_logs" 
            raw_logs = get_data(query_raw)
            all_projects = get_data("SELECT name FROM projects")['name'].tolist()
            col_f1, col_f2, col_f3 = st.columns(3)
            search_project = col_f1.multiselect("Projekt:", all_projects)
            all_workers = sorted(list(set(raw_logs['user_name'].unique()))) if not raw_logs.empty else []
            search_worker = col_f2.multiselect("Verantwortlich (Planer):", all_workers)
            
            if search_project:
                scaffolds_in_proj = get_data(f"SELECT s.number FROM scaffolds s JOIN projects p ON s.project_id = p.id WHERE p.name IN ({','.join(['?']*len(search_project))})", tuple(search_project))
                available_scaffolds = sorted(scaffolds_in_proj['number'].tolist())
            else:
                available_scaffolds = get_data("SELECT number FROM scaffolds")['number'].tolist()
            search_scaffold = col_f3.multiselect("Gerüst (Nr.):", available_scaffolds)

            query_master = '''
                SELECT 
                    p.name as Projekt,
                    s.number as Gerüstnummer,
                    s.description as Beschreibung,
                    s.acc as ACC,
                    s.volume_m3 as m3,
                    s.area_m2 as m2,
                    s.weight_to as "to",
                    s.material_cost as Materialwert,
                    w.user_name as Planer,
                    w.hours as Planungsstunden
                FROM scaffolds s
                JOIN projects p ON s.project_id = p.id
                LEFT JOIN work_logs w ON s.number = w.scaffold_number AND p.name = w.project_name
            '''
            master_raw = get_data(query_master)
            
            if not master_raw.empty:
                if search_project: master_raw = master_raw[master_raw['Projekt'].isin(search_project)]
                if search_worker: master_raw = master_raw[master_raw['Planer'].isin(search_worker)]
                if search_scaffold: master_raw = master_raw[master_raw['Gerüstnummer'].isin(search_scaffold)]

                agg_df = master_raw.groupby(['Projekt', 'Gerüstnummer', 'Beschreibung', 'ACC', 'm3', 'm2', 'to', 'Materialwert']).agg({
                    'Planungsstunden': 'sum',
                    'Planer': lambda x: ", ".join(sorted(list(set([str(i) for i in x if i is not None]))))
                }).reset_index()
                
                agg_df = agg_df.fillna(0)
                agg_df['Eur/to'] = agg_df.apply(lambda x: x['Materialwert'] / x['to'] if x['to'] > 0 else 0, axis=1)
                agg_df['Euro/m3'] = agg_df.apply(lambda x: x['Materialwert'] / x['m3'] if x['m3'] > 0 else 0, axis=1)
                agg_df['kg/m3'] = agg_df.apply(lambda x: (x['to'] * 1000) / x['m3'] if x['m3'] > 0 else 0, axis=1)

                final_columns_order = ['Gerüstnummer', 'm3', 'm2', 'to', 'Materialwert', 'Eur/to', 'Euro/m3', 'kg/m3', 'Planer', 'ACC', 'Beschreibung', 'Planungsstunden']
                display_cols = ['Projekt'] + [c for c in final_columns_order if c in agg_df.columns]
                final_df = agg_df[display_cols]

                st.dataframe(final_df, use_container_width=True, 
                             column_config={
                                 "m3": st.column_config.NumberColumn("m³", format="%.0f"),
                                 "m2": st.column_config.NumberColumn("m²", format="%.0f"),
                                 "to": st.column_config.NumberColumn("to", format="%.2f"),
                                 "Materialwert": st.column_config.NumberColumn("Materialwert", format="%.2f €"),
                                 "Eur/to": st.column_config.NumberColumn("Eur/to", format="%.2f €"),
                                 "Euro/m3": st.column_config.NumberColumn("Euro/m3", format="%.2f €"),
                                 "kg/m3": st.column_config.NumberColumn("kg/m3", format="%.2f kg"),
                                 "Planungsstunden": st.column_config.NumberColumn("Planungsstunden", format="%.1f h"),
                             })
                
                if not final_df.empty:
                    filename = get_export_filename(search_project)
                    st.download_button(label="📥 Master-Tabelle exportieren", data=to_excel(final_df, "Gerüste"), file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.divider()
            st.subheader("🛠 Stundenübersicht & Korrektur")
            query_details = "SELECT id, work_date as Datum, user_name as Name, scaffold_number as Gerüstnummer, hours as Stunden, comment as Anmerkungen, version as Versionsnummer, project_name as Projekt FROM work_logs WHERE 1=1"
            params_d = []
            if search_project:
                query_details += f" AND project_name IN ({','.join(['?']*len(search_project))})"
                params_d.extend(search_project)
            if search_worker:
                query_details += f" AND user_name IN ({','.join(['?']*len(search_worker))})"
                params_d.extend(search_worker)
            if search_scaffold:
                query_details += f" AND scaffold_number IN ({','.join(['?']*len(search_scaffold))})"
                params_d.extend(search_scaffold)
            query_details += " ORDER BY id DESC"
            df_details = get_data(query_details, tuple(params_d))
            st.dataframe(df_details, use_container_width=True, column_config={"id": None})
            
            if not df_details.empty:
                export_cols = ['Datum', 'Name', 'Gerüstnummer', 'Stunden', 'Anmerkungen', 'Versionsnummer']
                valid_export_cols = [c for c in export_cols if c in df_details.columns]
                df_export_stunden = df_details[valid_export_cols]
                filename_stunden = get_export_filename(search_project).replace("Engineering Stunden", "Stundenuebersicht")
                st.download_button(label="📥 Stundenübersicht exportieren", data=to_excel(df_export_stunden, "Stundenübersicht"), file_name=filename_stunden, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            with st.expander("Eintrag bearbeiten / löschen (Nach ID)", expanded=False):
                st.caption("Referenz-Tabelle (ID):")
                st.dataframe(df_details[['id', 'Datum', 'Name', 'Gerüstnummer', 'Stunden']], hide_index=True)
                edit_id_log = st.number_input("ID eingeben:", min_value=0, step=1)
                if edit_id_log > 0:
                    log_match = get_data("SELECT * FROM work_logs WHERE id=?", (edit_id_log,))
                    if not log_match.empty:
                        l_row = log_match.iloc[0]
                        with st.form("edit_log_adv"):
                            c1, c2 = st.columns(2)
                            new_comm = c1.text_input("Anmerkungen", value=l_row['comment'] if l_row['comment'] else "")
                            new_ver = c2.text_input("Version", value=l_row['version'] if l_row['version'] else "")
                            c3, c4 = st.columns(2)
                            new_h = c3.number_input("Stunden", value=float(l_row['hours']), step=0.5)
                            if st.form_submit_button("✅ Speichern"):
                                run_query("UPDATE work_logs SET hours=?, comment=?, version=? WHERE id=?", (new_h, new_comm, new_ver, edit_id_log))
                                st.success("Korrigiert!"); time.sleep(1); st.rerun()
                        if st.button("🗑️ Löschen"):
                            run_query("DELETE FROM work_logs WHERE id=?", (edit_id_log,))
                            st.rerun()

        # TAB 1: KPI
        with tab1:
            st.subheader("Projekt-Controlling")
            if not all_projects: st.warning("Keine Projekte.")
            else:
                col_kpi_1, col_kpi_2 = st.columns([1, 2])
                selected_project = col_kpi_1.selectbox("Projekt wählen", all_projects)
                scaffolds_data = get_data('SELECT s.volume_m3, s.material_cost FROM scaffolds s JOIN projects p ON s.project_id = p.id WHERE p.name = ?', (selected_project,))
                scaffolds_data = scaffolds_data.fillna(0)
                total_vol = scaffolds_data['volume_m3'].sum()
                total_cost = scaffolds_data['material_cost'].sum()
                avg_price = total_cost / total_vol if total_vol > 0 else 0
                st.markdown(f"**Projektkennzahlen: {selected_project}**")
                k1, k2, k3 = st.columns(3)
                k1.metric("Gesamtvolumen", f"{total_vol:,.0f} m³")
                k2.metric("Gesamtkosten (Materialwert)", f"{total_cost:,.0f} €")
                k3.metric("Ø Preis / m³", f"{avg_price:.2f} €")
                st.markdown("---")
                
                # DATA FETCH FOR CHARTS
                log_data = get_data("SELECT user_name, scaffold_number, hours FROM work_logs WHERE project_name = ?", (selected_project,))
                
                if not log_data.empty:
                    c_chart1, c_chart2 = st.columns(2)
                    
                    with c_chart1:
                        # Bar Chart: User Hours
                        bar_data = log_data.groupby('user_name')['hours'].sum().reset_index()
                        fig_bar = px.bar(bar_data, x='user_name', y='hours', title="Stunden pro Mitarbeiter", color='hours', color_continuous_scale='Blues')
                        st.plotly_chart(fig_bar, use_container_width=True)
                        
                    with c_chart2:
                        # Pie Chart: Scaffold Hours (Group small slices)
                        pie_data = log_data.groupby('scaffold_number')['hours'].sum().reset_index()
                        
                        # --- LOGIC FOR "SONSTIGE" ---
                        total_h = pie_data['hours'].sum()
                        if total_h > 0:
                            threshold = 0.03 * total_h # 3%
                            
                            main_data = pie_data[pie_data['hours'] >= threshold]
                            small_data = pie_data[pie_data['hours'] < threshold]
                            
                            if not small_data.empty:
                                other_sum = small_data['hours'].sum()
                                other_row = pd.DataFrame({'scaffold_number': ['Sonstige'], 'hours': [other_sum]})
                                pie_data_final = pd.concat([main_data, other_row], ignore_index=True)
                            else:
                                pie_data_final = main_data
                                
                            fig_pie = px.pie(pie_data_final, values='hours', names='scaffold_number', 
                                             title="Stundenverteilung (Top Gerüste)", hole=0.4)
                            st.plotly_chart(fig_pie, use_container_width=True)
                else:
                    st.info("Keine Arbeitsstunden für dieses Projekt gebucht.")

        # TAB 2: Workers
        with tab2:
            st.subheader("Personal")
            with st.form("aw"):
                n, p = st.text_input("Name"), st.text_input("Position")
                if st.form_submit_button("Hinzufügen"):
                    if n: run_query("INSERT INTO workers (name, position) VALUES (?, ?)", (n, p)); st.rerun()
            st.dataframe(get_data("SELECT * FROM workers"), hide_index=True, use_container_width=True)

        # TAB 3: Projects & Scaffolds
        with tab3:
            st.subheader("Stammdaten")
            c_p, c_s = st.columns(2)
            with c_p:
                st.markdown("### Projekte")
                with st.form("np"):
                    np = st.text_input("Name")
                    if st.form_submit_button("Erstellen"):
                        if np: run_query("INSERT INTO projects (name) VALUES (?)", (np,)); st.rerun()
                st.dataframe(get_data("SELECT * FROM projects"), hide_index=True)
            with c_s:
                st.markdown("### Gerüste (Übersicht)")
                pl = get_data("SELECT id, name FROM projects")
                if not pl.empty:
                    pm = dict(zip(pl['name'], pl['id']))
                    cp = st.selectbox("Projekt:", pl['name'], key="admin_scaf_view")
                    with st.expander("Gerüst manuell hinzufügen"):
                        with st.form("ns_admin"):
                            nn = st.text_input("Nummer")
                            nd = st.text_input("Beschreibung")
                            if st.form_submit_button("Speichern"):
                                run_query("INSERT INTO scaffolds (project_id, number, description, weight_to, material_cost) VALUES (?, ?, ?, 0, 0)", (pm[cp], nn, nd))
                                st.rerun()
                    sdf = get_data("SELECT number, description FROM scaffolds WHERE project_id=?", (pm[cp],))
                    st.dataframe(sdf, hide_index=True, use_container_width=True)

        # TAB 4: IMPORT (TRANSACTIONAL & SECURE)
        with tab4:
            st.subheader("📥 Excel-Import (Projekt-Zuordnung + Logs)")
            
            st.divider()
            
            # --- THE CUT (EXPANDER) ---
            with st.expander("🔧 Datenbank-Verwaltung & Logs (Geschützt)"):
                # Password Input
                db_password = st.text_input("Admin-Passwort eingeben:", type="password", key="db_pass_input")

                if db_password == "31337":
                    st.success("Zugriff gewährt!")
                    col_db1, col_db2 = st.columns(2)
                    
                    with col_db1:
                        st.markdown("### 🗑️ Datenbank löschen")
                        st.warning("Achtung: Dies löscht ALLE Projekte und Einträge!")
                        if st.button("🔴 KOMPLETT RESET (Tabellen löschen)", key="btn_reset"):
                            init_db(force_reset=True)
                            st.success("Datenbank wurde vollständig neu initialisiert!")
                            time.sleep(1)
                            st.rerun()

                    with col_db2:
                        st.markdown("### 📜 Import-Protokoll")
                        if 'import_logs' in st.session_state:
                            st.text_area("Logs", "\n".join(st.session_state['import_logs']), height=300)
                            if st.button("Logs leeren"):
                                del st.session_state['import_logs']
                                st.rerun()
                        else:
                            st.info("Keine Logs vorhanden.")
                elif db_password:
                    st.error("Falsches Passwort!")
            
            st.divider()
            st.markdown("Lädt Daten aus **'Gerüste'** und **'Stundenübersicht'**. \nProjekt wird aus dem **Dateinamen** ermittelt.")
            uploaded_file = st.file_uploader("Datei hochladen (z.B. 02-016_Projekt.xlsx)", type=['xlsx'])
            
            if uploaded_file:
                if st.button("Start Import"):
                    try:
                        filename = uploaded_file.name
                        match = re.match(r'^([\d-]+)', filename)
                        logs = []
                        
                        if match:
                            proj_prefix = match.group(1)
                            # Single Transaction
                            conn = sqlite3.connect(DB_FILE)
                            c = conn.cursor()
                            
                            # Check/Create Project
                            c.execute("SELECT id, name FROM projects WHERE name LIKE ?", (f"{proj_prefix}%",))
                            res = c.fetchone()
                            if res:
                                target_pid, target_pname = res
                                logs.append(f"✅ Projekt gefunden: {target_pname} (ID: {target_pid})")
                            else:
                                c.execute("INSERT INTO projects (name) VALUES (?)", (proj_prefix,))
                                target_pid = c.lastrowid
                                target_pname = proj_prefix
                                logs.append(f"🆕 Neues Projekt erstellt: {target_pname} (ID: {target_pid})")

                            xl = pd.ExcelFile(uploaded_file)
                            sheet_names = xl.sheet_names
                            
                            # 1. GERÜSTE
                            if 'Gerüste' in sheet_names:
                                logs.append("--- Tab 'Gerüste' ---")
                                df_scaf = pd.read_excel(uploaded_file, sheet_name='Gerüste')
                                df_scaf.columns = df_scaf.columns.str.strip()
                                
                                count_scaf = 0
                                for index, row in df_scaf.iterrows():
                                    s_num = clean_scaffold_number(row.get('Gerüstnummer', ''))
                                    s_m3 = safe_float(get_col_val(row, ['m3', 'm³', 'Volumen']))
                                    s_m2 = safe_float(get_col_val(row, ['m2', 'm²', 'Fläche']))
                                    s_to = safe_float(row.get('to'))
                                    s_mat = safe_float(row.get('Materialwert'))
                                    s_desc = str(row.get('Beschreibung', '')) if pd.notna(row.get('Beschreibung')) else ""
                                    s_acc = str(row.get('ACC', '')) if pd.notna(row.get('ACC')) else ""
                                    
                                    log_entry = f"Z.{index+2} [{s_num}]: "
                                    if s_num and s_num != 'nan':
                                        try:
                                            # Try INSERT
                                            c.execute("""
                                                INSERT INTO scaffolds (project_id, number, description, volume_m3, area_m2, weight_to, material_cost, acc) 
                                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                            """, (target_pid, s_num, s_desc, s_m3, s_m2, s_to, s_mat, s_acc))
                                            log_entry += "OK (NEU)"
                                            count_scaf += 1
                                        except sqlite3.IntegrityError:
                                            # If exists -> UPDATE
                                            c.execute("""
                                                UPDATE scaffolds SET description=?, volume_m3=?, area_m2=?, weight_to=?, material_cost=?, acc=? 
                                                WHERE project_id=? AND number=?
                                            """, (s_desc, s_m3, s_m2, s_to, s_mat, s_acc, target_pid, s_num))
                                            log_entry += "OK (UPDATED)"
                                            count_scaf += 1
                                    else: log_entry += "Ignoriert (Keine Nummer)"
                                    logs.append(log_entry)
                                logs.append(f"--> {count_scaf} Gerüste verarbeitet.")

                            # 2. STUNDEN
                            if 'Stundenübersicht' in sheet_names:
                                logs.append("\n--- Tab 'Stundenübersicht' ---")
                                df_hours = pd.read_excel(uploaded_file, sheet_name='Stundenübersicht')
                                df_hours.columns = df_hours.columns.str.strip()
                                count_hours, count_skip = 0, 0
                                for index, row in df_hours.iterrows():
                                    raw_date = row.get('Datum')
                                    if pd.isna(raw_date): w_date = date.today()
                                    else: w_date = pd.to_datetime(raw_date).date()
                                    u_name = str(row.get('Name', 'Importiert')).strip()
                                    s_num = clean_scaffold_number(row.get('Gerüstnummer', ''))
                                    h_val = parse_hours(row.get('Stunden'))
                                    comm = str(row.get('Anmerkungen', '')) if pd.notna(row.get('Anmerkungen')) else ""
                                    ver = str(row.get('Versionsnummer', '')) if pd.notna(row.get('Versionsnummer')) else ""
                                    
                                    if s_num:
                                        c.execute("SELECT id FROM work_logs WHERE user_name=? AND project_name=? AND scaffold_number=? AND work_date=? AND hours=? AND comment=? AND version=?", 
                                                           (u_name, target_pname, s_num, w_date, h_val, comm, ver))
                                        if not c.fetchone():
                                            c.execute("INSERT INTO work_logs (user_name, project_name, scaffold_number, work_date, hours, comment, version) VALUES (?, ?, ?, ?, ?, ?, ?)", 
                                                      (u_name, target_pname, s_num, w_date, h_val, comm, ver))
                                            count_hours += 1
                                        else: count_skip += 1
                                logs.append(f"--> {count_hours} Stunden importiert ({count_skip} Duplikate).")

                            conn.commit() # FINAL COMMIT
                            conn.close()
                            
                            st.session_state['import_logs'] = logs
                            st.rerun()
                            
                    except Exception as e:
                        if 'conn' in locals(): conn.close()
                        st.error(f"Kritischer Fehler: {e}")

            if 'import_logs' in st.session_state:
                with st.expander("🔍 Detaillierte Import-Logs", expanded=True):
                    st.text("\n".join(st.session_state['import_logs']))
                if st.button("Logs schließen"):
                    del st.session_state['import_logs']; st.rerun()

st.markdown("""<div class="footer"><p>Sergey Romanov, 2025 | Developed for promaintain®</p></div>""", unsafe_allow_html=True)